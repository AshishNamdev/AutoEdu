"""
Module: report

This module defines the `GenericReportExporter` class, which encapsulates
the logic for generating structured Excel and JSON reports from nested
record dictionaries keyed by unique identifiers.

Designed for AutoEdu and similar automation platforms, it supports
schema-independent flattening, audit-friendly exports, and automatic
backup of previous reports.

Features:
- Converts nested data into flat tabular format
- Saves both JSON and Excel versions of the report
- Backs up existing report files before overwriting
- Preserves input column order and appends audit columns at the end
- Cleans column labels for user-friendly exports
- Logs report generation status and file path

Usage:
    from report_exporter import GenericReportExporter

    exporter = GenericReportExporter(
        input_data,
        report_dir="reports/udise",
        filename="import_report"
    )
    exporter.save(first_column="Student PEN")

Author: Ashish Namdev (ashish28 [at] sirt [dot] gmail [dot] com)

Date Created:  2025-08-23
Last Modified: 2025-09-22

Version: 1.0.0
"""

import json
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from common.logger import logger
from utils import backup_file, clean_column_labels, clean_na_keys


class ReportExporter:
    """
    Generates structured Excel and JSON reports from nested record data.

    Features:
    - Backs up existing report files before overwriting
    - Preserves input column order and appends audit columns at the end
    - Cleans column labels for export
    - Schema-independent flattening of nested records
    """

    def __init__(self, input_data, report_sub_dir="udise", filename="report"):
        """
        Initializes the report exporter with data and configuration.

        Args:
            input_data (dict): Dictionary of records keyed by unique ID.
            report_dir (str): Directory under which reports are saved.
            filename (str): Base name for JSON and Excel files.
        """
        self.input_data = clean_na_keys(input_data)
        report_dir = os.path.join(os.getcwd(), "reports", report_sub_dir)
        self.report_json_file = os.path.join(report_dir, f"{filename}.json")
        self.report_excel_file = os.path.join(report_dir, f"{filename}.xlsx")

        os.makedirs(report_dir, exist_ok=True)

    def _backup_existing_reports(self):
        """
        Backs up existing report files (JSON and Excel) before overwriting.
        """
        for f in [self.report_json_file, self.report_excel_file]:
            if os.path.isfile(f):
                backup_file(f, logger)
                os.remove(f)

    def _flatten_input_data(self, first_column):
        """
        Flattens nested records into a list of row dictionaries.

        Args:
            first_column (str): Column name to assign the top-level key.

        Returns:
            List[Dict[str, Any]]: Flattened records.
        """
        rows = []
        for main_key, data in self.input_data.items():
            row = {first_column: main_key}
            if isinstance(data, dict):
                row.update(data)
                rows.append(row)
        return rows

    def _reorder_columns(self, df, audit_columns=None):
        """
        Ensures audit columns appear at the end of the DataFrame.

        Args:
            df (pd.DataFrame): Flattened DataFrame.
            audit_columns (list[str], optional): Columns to push to the end.

        Returns:
            pd.DataFrame: Reordered DataFrame.
        """
        audit_columns = audit_columns or ["Remark", "Import Status"]
        ordered_cols = [col for col in df.columns if col not in audit_columns] + [
            col for col in audit_columns if col in df.columns
        ]
        return df[ordered_cols]

    def save(self, first_column):
        """
        Generates and saves the report in both JSON and Excel formats.

        Args:
            first_column (str): Column name for the top-level key.

        Side Effects:
            - Writes to JSON and Excel files.
            - Logs output path.
        """
        self._backup_existing_reports()

        with open(self.report_json_file, "w", encoding="utf-8") as f:
            json.dump(self.input_data, f, indent=4, ensure_ascii=False)

        flattened_rows = self._flatten_input_data(first_column)
        df = pd.DataFrame(flattened_rows)
        df = self._reorder_columns(df)

        df.columns = [
            clean_column_labels(str(col), restore=True)
            for col in df.columns
        ]

        df.to_excel(self.report_excel_file, index=False)
        ReportStyler(self.report_excel_file).save()
        logger.info("Saved report to %s", self.report_excel_file)


class ReportStyler:
    """
    Applies professional formatting to Excel reports.

    Features:
    - Center-aligns all cells
    - Adds thin borders around all cells
    - Bolds and centers the header row
    - Optional freeze panes and column width adjustment
    """

    def __init__(self, filepath):
        """
        Args:
            filepath (str): Path to the Excel file to format.
        """
        self.filepath = filepath
        self.wb = load_workbook(filepath)
        self.ws = self.wb.active

        self.center = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        self.header_font = Font(bold=True)

    def apply_basic_formatting(self):
        """Applies borders, center alignment, and bold headers."""
        for row in self.ws.iter_rows():
            for cell in row:
                cell.alignment = self.center
                cell.border = self.thin_border
                if cell.row == 1:
                    cell.font = self.header_font

    def freeze_header(self):
        """Freezes the header row."""
        self.ws.freeze_panes = self.ws["A2"]

    def auto_adjust_column_widths(self):
        """Adjusts column widths based on content."""
        for col in self.ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col
            )
            col_letter = col[0].column_letter
            self.ws.column_dimensions[col_letter].width = max_length + 2

    def save(self):
        """
        Saves the formatted workbook.
        Overwrites the original file with applied styles.
        """
        self.apply_basic_formatting()
        self.freeze_header()
        self.auto_adjust_column_widths()
        self.wb.save(self.filepath)
